import io
from decimal import Decimal
from datetime import datetime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.shortcuts import get_object_or_404
from django.db.models import Q, F, Sum
from django.http import HttpResponse

from .models import Product, GRN, GRNItem, Sale, SaleItem, StockAdjustment, ProductImage, CosmeticOrder
from .serializers import (
    ProductSerializer, GRNSerializer, SaleSerializer, StockAdjustmentSerializer,
    ProductImageSerializer, CosmeticOrderSerializer,
)
from salons.models import Salon


def _get_salon_for_owner(request, salon_pk):
    salon = get_object_or_404(Salon, pk=salon_pk)
    if request.user.role == 'salon_owner' and salon.owner_id != request.user.id:
        return None, Response({'detail': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
    if request.user.role == 'client':
        return None, Response({'detail': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
    return salon, None


def _excel_response(headers, rows, sheet_title, filename):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='7C3AED')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _pdf_response(title, salon_name, headers, rows, filename):
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(salon_name, styles['Title']))
    elements.append(Paragraph(f"{title} — Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 16))

    data = [headers] + [[str(v) for v in row] for row in rows]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C3AED')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F3FF')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D1D5DB')),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)
    doc.build(elements)

    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


class AllCosmeticsView(APIView):
    """Returns all active, in-stock products from salons that have cosmetics enabled."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        products = (
            Product.objects
            .filter(is_active=True, current_stock__gt=0, salon__status='active', salon__cosmetics_enabled=True)
            .select_related('salon')
            .order_by('category', 'name')
        )
        data = []
        for p in products:
            data.append({
                'id': p.id,
                'salon': p.salon.id,
                'salon_name': p.salon.name,
                'name': p.name,
                'brand': p.brand,
                'category': p.category,
                'unit_of_measure': p.unit_of_measure,
                'selling_price': str(p.selling_price),
                'current_stock': p.current_stock,
            })
        return Response(data)


class SalonPublicCosmeticsView(APIView):
    """Public view: products for a specific salon (no auth required)."""
    permission_classes = [AllowAny]

    def get(self, request, salon_pk):
        salon = get_object_or_404(Salon, pk=salon_pk, status='active', cosmetics_enabled=True)
        qs = request.query_params.get('q', '').lower()
        cat = request.query_params.get('category', '')
        products = Product.objects.filter(salon=salon, is_active=True).order_by('category', 'name')
        if cat:
            products = products.filter(category=cat)
        if qs:
            products = products.filter(
                Q(name__icontains=qs) | Q(brand__icontains=qs) | Q(sku__icontains=qs)
            )
        data = []
        for p in products.prefetch_related('images'):
            first_img = p.images.first()
            first_image_url = (
                request.build_absolute_uri(first_img.image.url) if first_img and first_img.image else None
            )
            data.append({
                'id': p.id,
                'name': p.name,
                'brand': p.brand,
                'sku': p.sku,
                'category': p.category,
                'subcategory': p.subcategory,
                'shade_variant': p.shade_variant,
                'size': p.size,
                'unit_of_measure': p.unit_of_measure,
                'selling_price': str(p.selling_price),
                'current_stock': p.current_stock,
                'expiry_date': str(p.expiry_date) if p.expiry_date else None,
                'skin_type': p.skin_type,
                'status': p.status,
                'first_image_url': first_image_url,
            })
        return Response({'salon_name': salon.name, 'salon_id': salon.id, 'products': data})


class ProductSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        products = Product.objects.filter(salon=salon, is_active=True)
        total_skus = products.count()
        active_count = low_stock_count = out_stock_count = expiring_count = 0
        total_value = Decimal('0')
        for p in products:
            s = p.status
            if s == 'active':
                active_count += 1
            elif s == 'low_stock':
                low_stock_count += 1
            elif s == 'out_of_stock':
                out_stock_count += 1
            elif s == 'expiring_soon':
                expiring_count += 1
            total_value += p.selling_price * p.current_stock
        return Response({
            'total_skus': total_skus,
            'active': active_count,
            'low_stock': low_stock_count,
            'out_of_stock': out_stock_count,
            'expiring_soon': expiring_count,
            'total_value': str(total_value),
        })


class ProductListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        products = Product.objects.filter(salon=salon, is_active=True)
        return Response(ProductSerializer(products, many=True).data)

    def post(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(salon=salon)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk, pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        product = get_object_or_404(Product, pk=pk, salon=salon)
        return Response(ProductSerializer(product).data)

    def put(self, request, salon_pk, pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        product = get_object_or_404(Product, pk=pk, salon=salon)
        serializer = ProductSerializer(product, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, salon_pk, pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        product = get_object_or_404(Product, pk=pk, salon=salon)
        serializer = ProductSerializer(product, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, salon_pk, pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        product = get_object_or_404(Product, pk=pk, salon=salon)
        product.is_active = False
        product.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class GRNListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        grns = GRN.objects.filter(salon=salon).prefetch_related('items').order_by('-created_at')
        return Response(GRNSerializer(grns, many=True).data)

    def post(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        serializer = GRNSerializer(data=request.data)
        if serializer.is_valid():
            grn = serializer.save(salon=salon, created_by=request.user)
            return Response(GRNSerializer(grn).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class GRNConfirmView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, salon_pk, pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        grn = get_object_or_404(GRN, pk=pk, salon=salon)
        if grn.status == 'confirmed':
            return Response({'detail': 'GRN already confirmed'}, status=status.HTTP_400_BAD_REQUEST)
        for item in grn.items.all():
            item.product.current_stock += item.quantity_received
            item.product.save()
        grn.status = 'confirmed'
        grn.save()
        return Response(GRNSerializer(grn).data)


class SaleListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        sales = Sale.objects.filter(salon=salon).prefetch_related('items').order_by('-created_at')
        return Response(SaleSerializer(sales, many=True).data)

    def post(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        serializer = SaleSerializer(data=request.data, context={'salon_id': salon.id})
        if serializer.is_valid():
            sale = serializer.save(salon=salon, sold_by=request.user)
            # Check for low stock after sale and notify owner
            self._check_low_stock(sale, salon)
            return Response(SaleSerializer(sale).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def _check_low_stock(self, sale, salon):
        from users.utils import send_notification
        for item in sale.items.select_related('product'):
            p = item.product
            p.refresh_from_db()
            if p.current_stock <= p.reorder_level:
                send_notification(
                    salon.owner,
                    f"Low stock alert: {p.name} is at {p.current_stock} units (reorder level: {p.reorder_level}).",
                    notif_type='general',
                )


class StockAdjustmentListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        adjustments = StockAdjustment.objects.filter(salon=salon).order_by('-adjusted_at')
        return Response(StockAdjustmentSerializer(adjustments, many=True).data)

    def post(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        serializer = StockAdjustmentSerializer(data=request.data)
        if serializer.is_valid():
            product = serializer.validated_data['product']
            if product.salon_id != salon.id:
                return Response({'detail': 'Product does not belong to this salon'}, status=status.HTTP_400_BAD_REQUEST)
            adjustment = serializer.save(salon=salon, adjusted_by=request.user)
            product.current_stock += adjustment.quantity_change
            product.save()
            # Check low stock after adjustment
            product.refresh_from_db()
            if product.current_stock <= product.reorder_level:
                from users.utils import send_notification
                send_notification(
                    salon.owner,
                    f"Low stock alert: {product.name} is at {product.current_stock} units (reorder level: {product.reorder_level}).",
                    notif_type='general',
                )
            return Response(StockAdjustmentSerializer(adjustment).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class StockReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        products = Product.objects.filter(salon=salon, is_active=True)
        fmt = request.query_params.get('format', '').lower()

        if fmt == 'excel':
            headers = ['Name', 'Brand', 'Category', 'Unit', 'Cost Price', 'Selling Price', 'Stock', 'Reorder Level']
            rows = [[p.name, p.brand, p.category, p.unit_of_measure, p.cost_price, p.selling_price, p.current_stock, p.reorder_level] for p in products]
            return _excel_response(headers, rows, 'Stock Report', f'stock_report_{salon.id}.xlsx')

        if fmt == 'pdf':
            headers = ['Name', 'Brand', 'Category', 'Unit', 'Cost', 'Price', 'Stock', 'Reorder']
            rows = [[p.name, p.brand, p.category, p.unit_of_measure, p.cost_price, p.selling_price, p.current_stock, p.reorder_level] for p in products]
            return _pdf_response('Stock Report', salon.name, headers, rows, f'stock_report_{salon.id}.pdf')

        return Response(ProductSerializer(products, many=True).data)


class LowStockReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        products = Product.objects.filter(salon=salon, is_active=True, current_stock__lte=F('reorder_level'))
        fmt = request.query_params.get('format', '').lower()

        if fmt == 'excel':
            headers = ['Name', 'Brand', 'Category', 'Current Stock', 'Reorder Level', 'Shortage']
            rows = [[p.name, p.brand, p.category, p.current_stock, p.reorder_level, p.reorder_level - p.current_stock] for p in products]
            return _excel_response(headers, rows, 'Low Stock', f'low_stock_{salon.id}.xlsx')

        if fmt == 'pdf':
            headers = ['Name', 'Brand', 'Category', 'Stock', 'Reorder', 'Shortage']
            rows = [[p.name, p.brand, p.category, p.current_stock, p.reorder_level, p.reorder_level - p.current_stock] for p in products]
            return _pdf_response('Low Stock Report', salon.name, headers, rows, f'low_stock_{salon.id}.pdf')

        return Response(ProductSerializer(products, many=True).data)


class MovementsReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        from_date = request.query_params.get('from')
        to_date = request.query_params.get('to')
        fmt = request.query_params.get('format', '').lower()

        grn_qs = GRN.objects.filter(salon=salon, status='confirmed').prefetch_related('items__product')
        sale_qs = Sale.objects.filter(salon=salon).prefetch_related('items__product')
        adj_qs = StockAdjustment.objects.filter(salon=salon).select_related('product')

        if from_date:
            grn_qs = grn_qs.filter(created_at__date__gte=from_date)
            sale_qs = sale_qs.filter(created_at__date__gte=from_date)
            adj_qs = adj_qs.filter(adjusted_at__date__gte=from_date)
        if to_date:
            grn_qs = grn_qs.filter(created_at__date__lte=to_date)
            sale_qs = sale_qs.filter(created_at__date__lte=to_date)
            adj_qs = adj_qs.filter(adjusted_at__date__lte=to_date)

        if fmt == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook()

            # GRNs sheet
            ws1 = wb.active
            ws1.title = 'Stock Received'
            ws1.append(['GRN Ref', 'Supplier', 'Product', 'Qty', 'Unit Cost', 'Date'])
            for grn in grn_qs:
                for item in grn.items.all():
                    ws1.append([grn.reference_number, grn.supplier_name, item.product.name, item.quantity_received, float(item.unit_cost), grn.created_at.strftime('%Y-%m-%d')])

            # Sales sheet
            ws2 = wb.create_sheet('Sales')
            ws2.append(['Sale #', 'Product', 'Qty', 'Unit Price', 'Total', 'Date'])
            for sale in sale_qs:
                for item in sale.items.all():
                    ws2.append([sale.id, item.product.name, item.quantity, float(item.unit_price), float(item.quantity * item.unit_price), sale.created_at.strftime('%Y-%m-%d')])

            # Adjustments sheet
            ws3 = wb.create_sheet('Adjustments')
            ws3.append(['#', 'Product', 'Change', 'Reason', 'Notes', 'Date'])
            for adj in adj_qs:
                ws3.append([adj.id, adj.product.name, adj.quantity_change, adj.reason, adj.notes, adj.adjusted_at.strftime('%Y-%m-%d')])

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="movements_{salon.id}.xlsx"'
            return resp

        if fmt == 'pdf':
            grn_rows = []
            for grn in grn_qs:
                for item in grn.items.all():
                    grn_rows.append([grn.reference_number, grn.supplier_name, item.product.name, item.quantity_received, f"LKR {item.unit_cost}", grn.created_at.strftime('%Y-%m-%d')])
            sale_rows = []
            for sale in sale_qs:
                for item in sale.items.all():
                    sale_rows.append([sale.id, item.product.name, item.quantity, f"LKR {item.unit_price}", f"LKR {item.quantity * item.unit_price}", sale.created_at.strftime('%Y-%m-%d')])
            adj_rows = [[adj.id, adj.product.name, f"{adj.quantity_change:+d}", adj.reason, adj.adjusted_at.strftime('%Y-%m-%d')] for adj in adj_qs]

            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors

            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=landscape(letter))
            styles = getSampleStyleSheet()
            elements = []

            def make_table(title, headers, rows):
                elements.append(Paragraph(title, styles['Heading2']))
                elements.append(Spacer(1, 8))
                data = [headers] + [[str(v) for v in r] for r in rows] if rows else [headers, ['No data']]
                t = Table(data, repeatRows=1)
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7C3AED')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F3FF')]),
                    ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D1D5DB')),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 16))

            elements.append(Paragraph(salon.name, styles['Title']))
            elements.append(Paragraph(f"Movements Report — {datetime.now().strftime('%Y-%m-%d')}", styles['Normal']))
            elements.append(Spacer(1, 12))
            make_table('Stock Received (GRN)', ['GRN Ref', 'Supplier', 'Product', 'Qty', 'Unit Cost', 'Date'], grn_rows)
            make_table('Sales', ['Sale #', 'Product', 'Qty', 'Unit Price', 'Total', 'Date'], sale_rows)
            make_table('Adjustments', ['#', 'Product', 'Change', 'Reason', 'Date'], adj_rows)

            doc.build(elements)
            buf.seek(0)
            resp = HttpResponse(buf.read(), content_type='application/pdf')
            resp['Content-Disposition'] = f'attachment; filename="movements_{salon.id}.pdf"'
            return resp

        return Response({
            'grns': GRNSerializer(grn_qs, many=True).data,
            'sales': SaleSerializer(sale_qs, many=True).data,
            'adjustments': StockAdjustmentSerializer(adj_qs, many=True).data,
        })


class ProductImageListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, salon_pk, product_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        product = get_object_or_404(Product, pk=product_pk, salon=salon)
        images = product.images.all()
        return Response(ProductImageSerializer(images, many=True, context={'request': request}).data)

    def post(self, request, salon_pk, product_pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        product = get_object_or_404(Product, pk=product_pk, salon=salon)
        serializer = ProductImageSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save(product=product)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ProductImageDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, salon_pk, product_pk, pk):
        salon, err = _get_salon_for_owner(request, salon_pk)
        if err:
            return err
        product = get_object_or_404(Product, pk=product_pk, salon=salon)
        image = get_object_or_404(ProductImage, pk=pk, product=product)
        image.image.delete(save=False)
        image.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ProductPublicDetailView(APIView):
    """Public view: full product detail for a single product."""
    permission_classes = [AllowAny]

    def get(self, request, salon_pk, product_pk):
        salon = get_object_or_404(Salon, pk=salon_pk, status='active', cosmetics_enabled=True)
        product = get_object_or_404(Product, pk=product_pk, salon=salon, is_active=True)
        images = product.images.all()
        image_urls = [
            request.build_absolute_uri(img.image.url) for img in images if img.image
        ]
        return Response({
            'id': product.id,
            'salon_id': salon.id,
            'salon_name': salon.name,
            'name': product.name,
            'brand': product.brand,
            'sku': product.sku,
            'category': product.category,
            'subcategory': product.subcategory,
            'shade_variant': product.shade_variant,
            'size': product.size,
            'unit_of_measure': product.unit_of_measure,
            'cost_price': str(product.cost_price),
            'selling_price': str(product.selling_price),
            'current_stock': product.current_stock,
            'reorder_level': product.reorder_level,
            'supplier': product.supplier,
            'manufacturing_date': str(product.manufacturing_date) if product.manufacturing_date else None,
            'expiry_date': str(product.expiry_date) if product.expiry_date else None,
            'pao': product.pao,
            'barcode': product.barcode,
            'country_of_origin': product.country_of_origin,
            'certifications': product.certifications,
            'skin_type': product.skin_type,
            'notes': product.notes,
            'status': product.status,
            'images': image_urls,
        })


class ValidatePromoView(APIView):
    """Public: validate a promo code for a salon."""
    permission_classes = [AllowAny]

    def post(self, request, salon_pk):
        from datetime import date
        from salons.models import Offer
        from decimal import Decimal

        salon = get_object_or_404(Salon, pk=salon_pk, status='active')
        code = (request.data.get('code') or '').strip()
        subtotal = Decimal(str(request.data.get('subtotal', '0')))

        if not code:
            return Response({'valid': False, 'message': 'No code provided.'}, status=status.HTTP_400_BAD_REQUEST)

        today = date.today()
        offer = Offer.objects.filter(
            salon=salon, title__iexact=code, is_active=True,
            start_date__lte=today, end_date__gte=today,
        ).first()

        if not offer:
            return Response({'valid': False, 'message': 'Invalid or expired promo code.'})

        if offer.discount_type == 'percentage':
            discount = (subtotal * offer.discount_value / 100).quantize(Decimal('0.01'))
            label = f"{offer.discount_value}% off"
        else:
            discount = min(offer.discount_value, subtotal)
            label = f"LKR {offer.discount_value} off"

        return Response({
            'valid': True,
            'discount_type': offer.discount_type,
            'discount_value': str(offer.discount_value),
            'discount_amount': str(discount),
            'label': label,
            'description': offer.description,
        })


class CosmeticOrderListCreateView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, salon_pk):
        salon = get_object_or_404(Salon, pk=salon_pk, status='active', cosmetics_enabled=True)
        serializer = CosmeticOrderSerializer(data=request.data)
        if serializer.is_valid():
            order = serializer.save(
                salon=salon,
                client=request.user if request.user.role == 'client' else None,
            )
            return Response(CosmeticOrderSerializer(order).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def get(self, request, salon_pk):
        salon = get_object_or_404(Salon, pk=salon_pk, status='active')
        if request.user.role == 'client':
            orders = CosmeticOrder.objects.filter(salon=salon, client=request.user).prefetch_related('items').order_by('-created_at')
        else:
            salon, err = _get_salon_for_owner(request, salon_pk)
            if err:
                return err
            orders = CosmeticOrder.objects.filter(salon=salon).prefetch_related('items').order_by('-created_at')
        return Response(CosmeticOrderSerializer(orders, many=True).data)
